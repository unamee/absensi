import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils import timezone
from django.shortcuts import render
from django.template.loader import render_to_string
from django.http import HttpResponse
from .models import Attendance
from dept.models import Dept
from datetime import datetime, timedelta


def attendance_list(request):
    from_date_str = request.GET.get("from_date")
    to_date_str = request.GET.get("to_date")

    # default: tampilkan 100 record terakhir
    # report = Attendance.objects.select_related("employee", "machine").order_by("-timestamp")[:100]
    report = Attendance.objects.select_related(
        "employee__dept", "employee__jabatan", "machine"
    ).order_by("-timestamp")[:100]

    date_range = None

    if from_date_str and to_date_str:
        try:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
            report = (
                Attendance.objects.filter(timestamp__date__range=(from_date, to_date))
                .select_related("employee", "machine")
                .order_by("-timestamp")
            )
            date_range = (from_date.date(), to_date.date())
        except ValueError:
            pass  # jika format salah, biarkan default

    context = {
        "report": report,
        "from_date": from_date_str or "",
        "to_date": to_date_str or "",
        "date_range": date_range,
    }

    # ✅ Render partial jika via HTMX
    if request.headers.get("HX-Request") and "partial" in request.GET:
        return render(request, "attendance/partial/_attendance_table.html", context)

    # ✅ Render full page saat normal
    return render(request, "attendance/attendances_list.html", context)

def attendance_export_excel(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    # Ambil data sesuai filter
    qs = Attendance.objects.select_related(
        "employee__dept", "employee__jabatan", "machine"
    )

    if start_date and end_date:
        qs = qs.filter(timestamp__date__range=[start_date, end_date])
    else:
        today = timezone.now().date()
        qs = qs.filter(timestamp__date=today)

    # Buat workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Absensi"

    # Header Title
    ws.merge_cells("A1:H1")
    ws["A1"] = f"LAPORAN ABSENSI KARYAWAN ({start_date} s.d {end_date})"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Header table
    headers = [
        "No",
        "ID Karyawan",
        "Nama",
        "Departemen",
        "Jabatan",
        "Mesin",
        "Waktu Absen",
        "Status",
    ]
    ws.append(headers)

    # Style header
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for i, r in enumerate(qs.order_by("timestamp"), start=1):
        ws.append(
            [
                i,
                r.employee.id_karyawan if r.employee else "-",
                (
                    f"{r.employee.user.first_name} {r.employee.user.last_name}"
                    if r.employee
                    else "Tidak dikenal"
                ),
                r.employee.dept.nama_dept if r.employee and r.employee.dept else "-",
                (
                    r.employee.jabatan.nama_jabatan
                    if r.employee and r.employee.jabatan
                    else "-"
                ),
                r.machine.name if r.machine else "-",
                r.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                r.status or "-",
            ]
        )

    # Auto width
    # for column in ws.columns:
    #     max_length = 0
    #     column_letter = column[0].column_letter
    #     for cell in column:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(str(cell.value))
    #         except:
    #             pass
    #     adjusted_width = max_length + 2
    #     ws.column_dimensions[column_letter].width = adjusted_width

    from openpyxl.utils import get_column_letter

    for i, column_cells in enumerate(
        ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
        start=1,
    ):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    # Border
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Buat response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Laporan_Absensi_{start_date}_sampai_{end_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def daily_report(request):
    # ambil tanggal dari parameter GET atau gunakan hari ini
    date_str = request.GET.get("date")
    if date_str:
        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    else:
        selected_date = timezone.now().date()

    start = datetime.combine(selected_date, datetime.min.time())
    end = datetime.combine(selected_date, datetime.max.time())

    # ambil semua log absensi di tanggal tsb
    logs = Attendance.objects.filter(timestamp__range=(start, end)).select_related(
        "employee"
    )

    # kelompokkan berdasarkan employee
    report = []
    employees = Employee.objects.filter(
        id__in=[log.employee_id for log in logs if log.employee_id]
    )

    for emp in employees:
        emp_logs = [l for l in logs if l.employee_id == emp.id]
        if not emp_logs:
            continue

        # jam masuk = log pertama
        check_in = min(emp_logs, key=lambda l: l.timestamp).timestamp
        # jam keluar = log terakhir
        check_out = max(emp_logs, key=lambda l: l.timestamp).timestamp

        total_hours = (check_out - check_in).seconds / 3600

        report.append(
            {
                "id_karyawan": emp.id_karyawan,
                "nama": f"{emp.user.first_name} {emp.user.last_name}",
                "dept": emp.dept.nama if emp.dept else "-",
                "jabatan": emp.jabatan.nama if emp.jabatan else "-",
                "masuk": check_in.strftime("%H:%M:%S"),
                "pulang": check_out.strftime("%H:%M:%S"),
                "durasi": f"{total_hours:.2f} jam",
            }
        )

    context = {
        "date": selected_date,
        "report": report,
    }
    return render(request, "attendance/daily_report.html", context)


