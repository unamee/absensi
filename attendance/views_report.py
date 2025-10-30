from django.shortcuts import render
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from attendance.models import Attendance
from employee.models import Employee
from dept.models import Dept
from django.utils.timezone import localtime


def attendance_report(request):    
    depts = Dept.objects.all()
    start_str = request.GET.get("start_date")
    end_str = request.GET.get("end_date")
    selected_dept = request.GET.get("dept")

    if start_str and end_str:
        start_date = datetime.strptime(start_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_str, "%Y-%m-%d").date()
    else:
        today = timezone.now().date()
        start_date = today - timedelta(days=1)
        end_date = today

    # filter logs
    logs = Attendance.objects.filter(
        timestamp__date__range=(start_date, end_date)
    ).select_related("employee", "employee__dept", "employee__jabatan")

    print(selected_dept)

    # filter dept hanya jika valid
    if selected_dept and selected_dept.isdigit():
        logs = logs.filter(employee__dept_id=int(selected_dept))

    # struktur data report
    report = {}
    for log in logs:
        emp = log.employee
        if not emp:
            continue
        emp_id = emp.id_karyawan
        name = f"{emp.user.first_name} {emp.user.last_name}".strip()
        dept = emp.dept.nama_dept if emp.dept else "-"
        jabatan = emp.jabatan.nama_jabatan if emp.jabatan else "-"

        if emp_id not in report:
            report[emp_id] = {
                "nama": name,
                "dept": dept,
                "jabatan": jabatan,
                "harian": {},
            }

        local_ts = localtime(log.timestamp)
        tanggal = local_ts.date()
        jam = local_ts.strftime("%H:%M")

        if tanggal not in report[emp_id]["harian"]:
            report[emp_id]["harian"][tanggal] = {"masuk": jam, "pulang": jam}
        else:
            report[emp_id]["harian"][tanggal]["pulang"] = jam

    tanggal_list = list(next(iter(report.values()), {"harian": {}})["harian"].keys())

    context = {
        "depts": depts,
        "selected_dept": selected_dept,
        "report": report,
        "start_date": start_date,
        "end_date": end_date,
        "tanggal_list": tanggal_list,
    }

    # HTMX partial render
    if request.headers.get("HX-Request") and request.GET.get("partial") == "1":
        return render(request, "attendance/partial/_attendance_table_report.html", context)

    # Export Excel
    if request.GET.get("export") == "excel":
        return export_to_excel(report, start_date, end_date, selected_dept)

    return render(request, "attendance/attendance_report.html", context)



def export_to_excel(report, start_date, end_date, selected_dept=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Absensi"

    # Header perusahaan & periode
    ws.merge_cells("A1:I1")
    ws["A1"] = "PT WinnerSumbiri Knitting Factory"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Periode: {start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')}"
    ws["A2"].alignment = Alignment(horizontal="center")

    # Tambahkan nama departemen
    if selected_dept:
        try:
            dept_name = Dept.objects.get(id_dept=selected_dept).nama_dept
        except Dept.DoesNotExist:
            dept_name = "-"
        ws.merge_cells("A3:I3")
        ws["A3"] = f"Departemen: {dept_name}"
        ws["A3"].alignment = Alignment(horizontal="center")
    else:
        ws.merge_cells("A3:I3")
        ws["A3"] = "Departemen: Semua"
        ws["A3"].alignment = Alignment(horizontal="center")

    # Header kolom
    headers = ["No", "ID Karyawan", "Nama", "Dept", "Jabatan"]
    date_range = []
    current = start_date
    while current <= end_date:
        headers.append(current.strftime("%d-%b"))
        date_range.append(current)
        current += timedelta(days=1)
    headers.extend(["Total Hadir", "Total Alpha"])

    ws.append(headers)

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Isi data
    for i, (emp_id, data) in enumerate(report.items(), start=1):
        row = [i, emp_id, data["nama"], data["dept"], data["jabatan"]]
        hadir = 0
        alpha = 0
        for tgl in date_range:
            if tgl in data["harian"]:
                jam_masuk = data["harian"][tgl]["masuk"]
                jam_pulang = data["harian"][tgl]["pulang"]
                row.append(f"{jam_masuk}-{jam_pulang}")
                hadir += 1
            else:
                row.append("A")
                alpha += 1
        row.extend([hadir, alpha])
        ws.append(row)

    # Border & alignment
    for row in ws.iter_rows(min_row=5, max_col=len(headers)):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    # Simpan file response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"Laporan_Absensi_{start_date}_{end_date}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response
